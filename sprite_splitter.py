import json
import os
import argparse
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

def get_frames(data):
    """Extract frames data from different JSON structures."""
    if 'textures' in data:
        return data['textures'][0]['frames']
    elif 'frames' in data:
        return data['frames']
    else:
        raise ValueError("Unsupported JSON structure. Cannot find frames data.")

def split_sprite(json_file, sprite_image, wb, ws):
    # Load JSON data
    with open(json_file, 'r') as f:
        data = json.load(f)

    # Open the sprite image
    sprite = Image.open(sprite_image)

    # Create output directory if it doesn't exist
    output_dir = 'output_sheet'
    os.makedirs(output_dir, exist_ok=True)

    try:
        # Extract frame information from JSON
        frames = get_frames(data)

        # Process only the first frame
        if isinstance(frames, dict):
            frame_data = next(iter(frames.values()))
        elif isinstance(frames, list):
            frame_data = frames[0]
        else:
            raise ValueError(f"Unsupported frames data type: {type(frames)}")

        if isinstance(frame_data, dict):
            if 'frame' in frame_data:
                # TexturePacker format
                x = frame_data['frame']['x']
                y = frame_data['frame']['y']
                w = frame_data['frame']['w']
                h = frame_data['frame']['h']
            else:
                # Simple dictionary format
                x = frame_data['x']
                y = frame_data['y']
                w = frame_data['w']
                h = frame_data['h']
        elif isinstance(frame_data, list):
            # List format
            x, y, w, h = frame_data
        else:
            raise ValueError(f"Unsupported frame data format: {type(frame_data)}")

        # Extract the individual image from the sprite
        single_image = sprite.crop((x, y, x + w, y + h))

        # Generate the new filename
        base_name = os.path.splitext(os.path.basename(sprite_image))[0]
        filename = f"{base_name}.png"

        # Save the image
        output_path = os.path.join(output_dir, filename)
        single_image.save(output_path)
        print(f"Saved {output_path}")

        # Add to spreadsheet
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=filename)
        img = XLImage(output_path)
        img.width = 100  # Adjust as needed
        img.height = 100  # Adjust as needed
        ws.add_image(img, f'B{row}')

    except ValueError as e:
        print(f"Error processing {json_file}: {str(e)}")
    except KeyError as e:
        print(f"Error processing {json_file}: Missing key {str(e)}")

def process_directory(directory):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sprite Images"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # Get all JSON files in the directory
    json_files = [f for f in os.listdir(directory) if f.endswith('.json')]

    for json_file in json_files:
        # Construct the full path to the JSON file
        json_path = os.path.join(directory, json_file)

        # Construct the corresponding PNG file name
        png_file = os.path.splitext(json_file)[0] + '.png'
        png_path = os.path.join(directory, png_file)

        # Check if the corresponding PNG file exists
        if os.path.exists(png_path):
            print(f"Processing {json_file} and {png_file}")
            split_sprite(json_path, png_path, wb, ws)
        else:
            print(f"Warning: No matching PNG file found for {json_file}")

    # Save the workbook
    wb.save("sprite_images.xlsx")
    print("Spreadsheet saved as sprite_images.xlsx")

if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(description="Split sprite sheets and create a spreadsheet of images.")
    parser.add_argument("input_dir", nargs="?", default=os.getcwd(),
                        help="Directory containing JSON and PNG files (default: current directory)")

    # Parse arguments
    args = parser.parse_args()

    # Use the specified input directory or default to current directory
    input_directory = os.path.abspath(args.input_dir)

    print(f"Processing files in: {input_directory}")
    
    # Process all files in the input directory
    process_directory(input_directory)